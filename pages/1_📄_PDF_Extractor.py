import streamlit as st
import pdfplumber
import pandas as pd
import io
import os
import re
import openpyxl
import copy

# ==========================================
# CONFIGURAZIONE PAGINA E STATO
# ==========================================
st.set_page_config(page_title="PDF Extractor", page_icon="📄", layout="centered")

if "market_choice" not in st.session_state:
    st.session_state.market_choice = "Seleziona un'opzione..."
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

def reset_app_state():
    st.session_state.market_choice = "Seleziona un'opzione..."
    st.session_state.uploader_key += 1

# ==========================================
# STILI CUSTOM E INTERFACCIA
# ==========================================
color_primary = "#2A85C4"
color_text = "#333333"

st.markdown(f"""
    <style>
        .stApp {{ background-color: #F8F9FA; }}
        h1, h2, h3, p {{ color: {color_text} !important; }}
        .stFileUploader {{ border: 1px solid #E0E0E0; border-radius: 10px; background-color: #FFFFFF; padding: 20px; margin-bottom: 15px; }}
        .dataframe thead th {{ background-color: {color_text} !important; color: white !important; font-weight: bold; }}
        div.stDownloadButton > button:first-child {{
            background-color: {color_primary} !important; color: white !important;
            width: 100%; height: 60px; font-size: 1.2rem; font-weight: bold; border-radius: 10px; border: none; box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        div.stDownloadButton > button:first-child:hover {{ background-color: #1e69a0 !important; }}
    </style>
""", unsafe_allow_html=True)

# ==========================================
# ESPRESSIONI REGOLARI E COSTANTI
# ==========================================
# MODIFICA: Supporto per la data "Presente"
REGEX_DATE = re.compile(r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}\s*[-\u2013\u2014a-zA-Z]+\s*(?:\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|Presente))', re.IGNORECASE)
REGEX_VAL_PIPE = re.compile(r'^\|\s*([+-]?\s*\d[\d\s\.,]*(?:€|EUR)?\s*\.?)$')
REGEX_VAL_NOPIPE = re.compile(r'^([+-]?\s*\d[\d\s\.,]*(?:€|EUR)?\s*\.?)$')
# MODIFICA: La regex FULL ora tollera le virgole come separatori
REGEX_VAL_FULL = re.compile(r'^([^\|]+?)(?:\s+|\|\s*|,\s*)([+-]?\s*\d[\d\s\.,]*(?:€|EUR)?\s*\.?)$')

SECTIONS = ["Ordini standard", "Saldo Iniziale", "Vendite", "Rimborsi", "Spese", "Importo Riserva dell'Account", "Ricavi Netti"]

STOP_WORDS = {'di', 'a', 'da', 'in', 'con', 'su', 'per', 'tra', 'fra', 'il', 'lo', 'la', 'i', 'gli', 'le', 'un', 'uno', 'una', 'ed', 'e', 
              'nell', 'dell', 'sull', 'all', 'al', 'allo', 'alla', 'ai', 'agli', 'alle', 'del', 'dello', 'della', 'dei', 'degli', 'delle', 
              'nel', 'nello', 'nella', 'nei', 'negli', 'nelle'}

# ==========================================
# FUNZIONI DI SUPPORTO (Normalizzazione)
# ==========================================
def clean_value(val_str):
    if not val_str: return None
    cleaned = re.sub(r'[^\d,.-]', '', val_str).rstrip('.-,')
    if not cleaned: return None
    if ',' in cleaned and '.' in cleaned:
        if cleaned.rfind(',') > cleaned.rfind('.'): cleaned = cleaned.replace('.', '').replace(',', '.')
        else: cleaned = cleaned.replace(',', '')
    elif ',' in cleaned: cleaned = cleaned.replace(',', '.')
    try: return float(cleaned)
    except ValueError: return None

def normalize_name(name):
    """Normalizza i nomi ignorando preposizioni e casi critici come la riserva."""
    if not name: return ""
    name_lower = str(name).lower()
    if "riserva" in name_lower: return "riserva account" # Blindatura forte per voce riserva
    name_clean = name_lower.replace("'", " ").replace("’", " ")
    words = re.split(r'\W+', name_clean)
    return " ".join([w for w in words if w and w not in STOP_WORDS])

def generate_id(nome, sezione, seen_refs):
    sec_initial = sezione.strip()[0].upper() if sezione else "X"
    words = [w for w in re.split(r'\W+', nome) if w and w.lower() not in STOP_WORDS]
    
    if len(words) >= 2: prefix = words[0][:2].capitalize().ljust(2, 'a') + words[1][:2].capitalize().ljust(2, 'a')
    elif len(words) == 1: prefix = words[0][:4].capitalize().ljust(4, 'a')
    else: prefix = "Xxxx"
    
    prefix = prefix[:4]
    base_rif = f"{prefix}_{sec_initial}"
    rif = base_rif
    counter = 2
    
    while rif in seen_refs:
        rif = f"{prefix}{counter}_{sec_initial}"
        counter += 1
        
    seen_refs.add(rif)
    return rif

def sanitize_filename(filename):
    return re.sub(r'[/\\:*?"<>|]', '-', filename)

# ==========================================
# MOTORE DI CLONAZIONE ANAGRAFICA 
# ==========================================
def load_mapping():
    mapping = {}
    seen = set()
    headers = ["ID", "Nome", "Sezione"]
    idx_nome, idx_sez, idx_id = 1, 2, 0
    count = 0
    
    files = ["Tabella ID_5.xlsx", "Tabella ID_4.xlsx", "Tabella ID_3.xlsx", "Tabella ID.xlsx - Foglio1.csv", "Tabella ID.csv", "Tabella ID.xlsx"]
    found = next((f for f in files if os.path.exists(f)), None)

    if not found:
        return mapping, seen, headers, idx_nome, idx_sez, idx_id, False, "Nessun file anagrafica trovato"

    try:
        if found.endswith('.xlsx'):
            wb = openpyxl.load_workbook(found, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else f"Colonna_{i}" for i, cell in enumerate(ws[1])]
            hl = [h.lower() for h in headers]
            idx_nome = hl.index('nome') if 'nome' in hl else -1
            idx_sez = hl.index('sezione') if 'sezione' in hl else -1
            idx_id = hl.index('id') if 'id' in hl else -1
            
            for row in ws.iter_rows(min_row=2):
                if idx_nome != -1 and idx_sez != -1 and row[idx_nome].value and row[idx_sez].value:
                    n_key = normalize_name(row[idx_nome].value)
                    s_key = str(row[idx_sez].value).strip().lower()
                    i_val = str(row[idx_id].value).strip() if idx_id != -1 else ""
                    
                    row_styles = []
                    for cell in row:
                        row_styles.append({
                            'fill': copy.copy(cell.fill) if hasattr(cell, 'fill') and cell.fill else None,
                            'font': copy.copy(cell.font) if hasattr(cell, 'font') and cell.font else None,
                            'alignment': copy.copy(cell.alignment) if hasattr(cell, 'alignment') and cell.alignment else None,
                            'border': copy.copy(cell.border) if hasattr(cell, 'border') and cell.border else None
                        })

                    key = (n_key, s_key)
                    if key not in mapping: mapping[key] = []
                    mapping[key].append({'vals': [c.value for c in row], 'styles': row_styles, 'id': i_val})
                    if i_val: seen.add(i_val)
                    count += 1
        else:
            df = pd.read_csv(found)
            headers = list(df.columns)
            hl = [h.lower() for h in headers]
            idx_nome = hl.index('nome') if 'nome' in hl else -1
            idx_sez = hl.index('sezione') if 'sezione' in hl else -1
            idx_id = hl.index('id') if 'id' in hl else -1
            
            for _, row in df.iterrows():
                if idx_nome != -1 and idx_sez != -1 and pd.notna(row.iloc[idx_nome]) and pd.notna(row.iloc[idx_sez]):
                    n_key = normalize_name(row.iloc[idx_nome])
                    s_key = str(row.iloc[idx_sez]).strip().lower()
                    i_val = str(row.iloc[idx_id]).strip() if idx_id != -1 else ""
                    
                    key = (n_key, s_key)
                    if key not in mapping: mapping[key] = []
                    mapping[key].append({'vals': list(row), 'styles': [{'fill': None, 'font': None, 'alignment': None, 'border': None} for _ in range(len(row))], 'id': i_val})
                    if i_val: seen.add(i_val)
                    count += 1
        return mapping, seen, headers, idx_nome, idx_sez, idx_id, True, f"Caricati {count} ID da {found}"
    except Exception as e:
        return mapping, seen, headers, idx_nome, idx_sez, idx_id, False, str(e)

# ==========================================
# MOTORE DI ESTRAZIONE PDF
# ==========================================
def extract_pdf_data(pdf_file, mapping, seen, headers, idx_nome, idx_sez, idx_id):
    final_data = []
    new_items = []
    extracted_period = "Periodo_Non_Rilevato"
    current_sec, started, finished = None, False, False

    def create_blank_item(nome, val, sec, is_date=False):
        row_id = generate_id(nome, sec, seen)
        vals = [None] * len(headers)
        if idx_id != -1: vals[idx_id] = row_id
        if idx_nome != -1: vals[idx_nome] = nome
        if idx_sez != -1: vals[idx_sez] = sec
        new_items.append({"ID": row_id, "Nome": nome, "Valore": val, "Sezione": sec})
        return {'vals': vals, 'styles': [{'fill': None, 'font': None, 'alignment': None, 'border': None} for _ in range(len(headers))]}

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text(layout=True) or page.extract_text()
            if not text or finished: continue
            
            pending_name = ""
            for line in text.split('\n'):
                # MODIFICA: Pulizia dei caratteri CSV "fantasma" che corrompono le espressioni regolari
                line = line.strip().replace('"', '')
                if line.startswith(','):
                    line = line[1:].strip()

                if not line or finished: continue

                if not started:
                    if "ordini standard" in line.lower():
                        started = True
                        current_sec = "Ordini standard"
                        match = REGEX_DATE.search(line)
                        if match:
                            raw_date = match.group(1).strip()
                            d_parts = re.split(r'\s*[-\u2013\u2014a-zA-Z]+\s*', raw_date)
                            if len(d_parts) == 2:
                                d1 = d_parts[0].replace('/', '.').replace('-', '.')
                                d2 = d_parts[1].replace('/', '.').replace('-', '.')
                                extracted_period = f"{d1}-{d2}"
                            else:
                                extracted_period = sanitize_filename(raw_date)
                        else:
                            extracted_period = sanitize_filename(re.split(r'(?i)ordini standard', line)[-1].strip() or "Data non rilevata")
                        
                        key = (normalize_name("Ordini standard"), "ordini standard")
                        item = mapping[key].pop(0).copy() if mapping.get(key) else create_blank_item("Ordini standard", extracted_period, "Ordini standard", True)
                        final_data.append({'row': item['vals'], 'styles': item['styles'], 'valore_estratto': extracted_period, 'sezione': 'Ordini standard', 'nome': 'Ordini standard'})
                    continue 

                nome, val_grezzo = None, None
                if match_pipe := REGEX_VAL_PIPE.search(line):
                    if pending_name: nome, val_grezzo = pending_name.strip(), match_pipe.group(1).strip(); pending_name = ""
                    else: continue
                elif match_nopipe := REGEX_VAL_NOPIPE.search(line):
                    if pending_name: nome, val_grezzo = pending_name.strip(), match_nopipe.group(1).strip(); pending_name = ""
                    else: continue
                elif match_full := REGEX_VAL_FULL.search(line):
                    nome, val_grezzo = match_full.group(1).strip(), match_full.group(2).strip(); pending_name = ""
                else:
                    clean_l = line.replace('|', '').strip()
                    lower_l = clean_l.lower()
                    if "saldo iniziale" in lower_l: clean_l = "Saldo Iniziale"
                    elif "ricavi netti" in lower_l: clean_l = "Ricavi Netti"
                    elif "importo riserva" in lower_l or "codice di riserva" in lower_l: clean_l = "Importo Riserva dell'Account"
                    
                    if clean_l.lower() in [s.lower() for s in SECTIONS]: pending_name = clean_l
                    else: pending_name = (pending_name + " " + clean_l if pending_name else clean_l)[:250]
                    continue

                if nome.endswith('-'): nome, val_grezzo = nome[:-1].strip(), '-' + val_grezzo
                val_num = clean_value(val_grezzo)
                
                if val_num is not None and nome:
                    n_lower = nome.lower()
                    if "saldo iniziale" in n_lower: nome = "Saldo Iniziale"
                    elif "ricavi netti" in n_lower: nome = "Ricavi Netti"
                    elif "importo riserva" in n_lower or "codice di riserva" in n_lower: nome = "Importo Riserva dell'Account"

                    for sec in SECTIONS:
                        if nome.lower() == sec.lower(): current_sec = sec; break
                    
                    n_key = normalize_name(nome)
                    key = (n_key, current_sec.lower() if current_sec else "")
                    item = mapping[key].pop(0).copy() if mapping.get(key) else create_blank_item(nome, val_num, current_sec)
                    
                    final_data.append({'row': item['vals'], 'styles': item['styles'], 'valore_estratto': val_num, 'sezione': current_sec, 'nome': nome})
                    
                    if current_sec == "Ricavi Netti" and "ricavi netti" in n_lower: finished = True; break
                            
    return final_data, pd.DataFrame(new_items), extracted_period

# ==========================================
# INIETTORE EXCEL STILIZZATO
# ==========================================
def create_styled_excel(final_data, headers, idx_nome):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report_Estratto"
    
    ins_pos = idx_nome + 1 if idx_nome != -1 else len(headers)
    out_headers = headers.copy()
    out_headers.insert(ins_pos, "Valore Estratto PDF")
    ws.append(out_headers)
    
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    
    for item in final_data:
        row = list(item['row'])
        styles = list(item['styles'])
        
        row.insert(ins_pos, item['valore_estratto'])
        styles.insert(ins_pos, {'fill': None, 'font': None, 'alignment': None, 'border': None})
        ws.append(row)
        
        current_row = ws.max_row
        
        for col_num, style in enumerate(styles, start=1):
            cell = ws.cell(row=current_row, column=col_num)
            if style['fill']:
                try: cell.fill = style['fill']
                except: pass
            if style['font']:
                try: cell.font = style['font']
                except: pass
            if style['alignment']:
                try: cell.alignment = style['alignment']
                except: pass
            if style['border']:
                try: cell.border = style['border']
                except: pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf

# ==========================================
# BOX DI QUADRATURA
# ==========================================
def render_quadratura_box(final_data, headers):
    st.markdown("<div style='background-color: #EBF5FB; padding: 15px; border-radius: 10px; border: 1px solid #AED6F1; margin-bottom: 20px;'>", unsafe_allow_html=True)
    st.markdown("<h4 style='color: #2A85C4; margin-top:0px;'>📊 Box di Quadratura Fiscale</h4>", unsafe_allow_html=True)
    
    hl = [h.lower() for h in headers]
    idx_p = hl.index('principale') if 'principale' in hl else -1
    idx_s = hl.index('secondaria') if 'secondaria' in hl else -1
    idx_f = hl.index('figlia') if 'figlia' in hl else -1
    
    for sec_name in ["Vendite", "Rimborsi", "Spese"]:
        v_princ, v_sec, has_data = 0.0, 0.0, False
        
        for item in final_data:
            if item.get('sezione') != sec_name: continue
            val = item['valore_estratto']
            if not isinstance(val, (int, float)): continue
            has_data = True
            
            row = item['row']
            is_p = (idx_p != -1 and str(row[idx_p]).strip() not in ["", "nan", "None"])
            is_s = (idx_s != -1 and str(row[idx_s]).strip() not in ["", "nan", "None"])
            is_f = (idx_f != -1 and str(row[idx_f]).strip() not in ["", "nan", "None"])
            
            if not (is_p or is_s or is_f):
                is_p = (item['nome'].lower() == sec_name.lower())
                is_s = not is_p
            
            if is_p: v_princ += val
            elif is_s and not is_f: v_sec += val
            
        if has_data:
            diff = round(abs(v_princ - v_sec), 2)
            if diff <= 0.05: st.markdown(f"✅ **{sec_name}**: Quadratura perfetta! (Principale: {v_princ:,.2f} € | Secondarie: {v_sec:,.2f} €)")
            else: st.markdown(f"❌ <span style='color:red;'>**{sec_name}**: Disallineamento! (Principale: {v_princ:,.2f} € | Secondarie: {v_sec:,.2f} € | Diff: {diff:,.2f} €)</span>", unsafe_allow_html=True)
    
    st.markdown("</div>", unsafe_allow_html=True)

# ==========================================
# ESECUZIONE APP (UI Layout)
# ==========================================
logo_path = "logo.png"
if os.path.exists(logo_path):
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2: st.image(logo_path, use_container_width=True)
st.markdown("<h2 style='text-align: center; font-weight: bold;'>PDF to Excel Extractor</h2>", unsafe_allow_html=True)

st.markdown("<h4 style='color: #2A85C4;'>1. Verifica Anagrafica ID</h4>", unsafe_allow_html=True)
map_dict, seen, headers, idx_n, idx_s, idx_id, map_success, map_msg = load_mapping()

if map_success: st.success(f"✅ {map_msg}")
else: st.warning(f"⚠️ {map_msg}. Verranno usate colonne standard.")

st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("<h4 style='color: #2A85C4;'>2. Scegli il Marketplace e Carica PDF</h4>", unsafe_allow_html=True)

marketplaces = {
    "Seleziona un'opzione...": None,
    "🇮🇹 IT - Italia": "AM_IT", "🇫🇷 FR - Francia": "AM_FR", "🇩🇪 DE - Germania": "AM_DE",
    "🇪🇸 ES - Spagna": "AM_ES", "🇳🇱 NL - Olanda": "AM_NL", "🇵🇱 PL - Polonia": "AM_PL",
    "🇸🇪 SE - Svezia": "AM_SE", "🇮🇪 IE - Irlanda": "AM_IE", "🌍 XX - Altro": "AM_XX"
}

sel_market = st.selectbox("Marketplace:", list(marketplaces.keys()), key="market_choice")
m_prefix = marketplaces[sel_market]

pdf_file = st.file_uploader("", type="pdf", key=f"pdf_up_{st.session_state.uploader_key}")

if pdf_file:
    if not m_prefix:
        st.error("⚠️ ATTENZIONE: Seleziona il Paese prima di procedere!")
    else:
        with st.spinner('Estrazione e clonazione stilistica in corso...'):
            final_data, df_new, ext_period = extract_pdf_data(pdf_file, map_dict, seen, headers, idx_n, idx_s, idx_id)
            
            if final_data:
                st.success("Estrazione completata!")
                render_quadratura_box(final_data, headers)
                
                # ESTRAZIONE NUMERO FILE SORGENTE
                file_num = ""
                num_match = re.search(r'AM_[A-Z]{2}\s*(\d+)', pdf_file.name, re.IGNORECASE)
                if num_match:
                    file_num = f" {num_match.group(1)}"
                
                # NOME DEL FILE OUTPUT AGGIORNATO COME RICHIESTO
                exp_file = f"{m_prefix}{file_num} - Estratto Conto Tab {ext_period}.xlsx"
                
                if not df_new.empty:
                    st.warning(f"⚠️ Generate {len(df_new)} NUOVE voci.")
                    st.dataframe(df_new, use_container_width=True)
                    buf_new = io.BytesIO()
                    with pd.ExcelWriter(buf_new, engine='openpyxl') as w: df_new.to_excel(w, index=False, sheet_name='Nuovi_ID')
                    st.download_button("SCARICA NUOVE VOCI 📥", buf_new.getvalue(), f"{m_prefix}_nuovi_id_{ext_period}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else: st.info("Tutte le voci riconosciute.")
                
                st.markdown("<p style='font-size:14px; color:#666;'><em>💡 <b>Nota Salvataggio:</b> Il file viene scaricato automaticamente. Se vuoi scegliere la cartella di destinazione, assicurati di aver attivato l'opzione 'Chiedi dove salvare il file prima di scaricarlo' nelle impostazioni del tuo browser web.</em></p>", unsafe_allow_html=True)
                
                excel_buf = create_styled_excel(final_data, headers, idx_n)
                st.download_button(f"SCARICA ESTRATTO CONTO TAB 📥", excel_buf.getvalue(), exp_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", on_click=reset_app_state)
            else: st.error("Nessun dato trovato nel PDF.")
