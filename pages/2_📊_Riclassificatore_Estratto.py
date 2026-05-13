import streamlit as st
import pandas as pd
from decimal import Decimal, ROUND_HALF_UP
from PIL import Image
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
import io
import re

# --- 1. CONFIGURAZIONE PAGINA E STILE ---
st.set_page_config(page_title="Binky - Calcolatore Corrispettivi", page_icon="📊", layout="wide")

st.markdown("""
<style>
.stButton>button { background-color: #1A7EBF; color: white; border-radius: 8px; border: none; padding: 10px 24px; font-weight: bold; width: 100%; }
.stButton>button:hover { background-color: #145e91; color: white; }
h1, h2, h3 { color: #5C5D5F; }
.metric-box { background-color: #f8f9fa; border-left: 5px solid #1A7EBF; padding: 20px; border-radius: 8px; margin-top: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); height: 100%; }
.metric-title { font-size: 16px; color: #6c757d; font-weight: bold; margin-bottom: 5px; text-transform: uppercase; }
.metric-value { font-size: 28px; color: #2A85C4; font-weight: bold; margin: 0; }
.metric-sub { font-size: 14px; color: #5C5D5F; margin-top: 15px; padding-top: 15px; border-top: 1px solid #dee2e6; line-height: 1.6; }
.period-box { background-color: #E8F4F8; border-left: 5px solid #2A85C4; padding: 12px 20px; border-radius: 8px; margin-bottom: 20px; font-size: 16px; color: #145e91; font-weight: 500; display: flex; align-items: center; gap: 10px;}
</style>
""", unsafe_allow_html=True)

# --- 2. LOGO E INTESTAZIONE ---
logo_path = "full_logo.jpg"
if os.path.exists(logo_path):
    st.image(Image.open(logo_path), width=250)

st.title("Calcolatore Corrispettivi Amazon")
st.markdown("Generazione automatica da Report Estratto")

# --- 3. DIZIONARI E FUNZIONI DI SUPPORTO ---
ALIQUOTE_IVA = {
    "AM_IT": Decimal('0.22'), "AM_FR": Decimal('0.20'), "AM_DE": Decimal('0.19'),
    "AM_ES": Decimal('0.21'), "AM_NL": Decimal('0.21'), "AM_SE": Decimal('0.25'),
    "AM_IE": Decimal('0.23'), "AM_XX": Decimal('0.00')
}

def dec(v):
    """Converte un valore in formato Decimal, con arrotondamento finanziario sicuro."""
    try:
        if pd.isna(v) or str(v).strip() == '': return Decimal('0.00')
        return Decimal(str(v)).quantize(Decimal('0.01'), ROUND_HALF_UP)
    except:
        return Decimal('0.00')

def estrai_colori_tema(file_buffer):
    """Legge il file Excel una sola volta e mappa i colori di sfondo degli ID Estratto."""
    mappa_colori = {}
    try:
        file_buffer.seek(0)
        wb = openpyxl.load_workbook(file_buffer, data_only=True)
        ws = wb.active
        header = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        idx_id_est = header.index('ID Estratto') if 'ID Estratto' in header else -1
        idx_voce = header.index('Voce Estratto') if 'Voce Estratto' in header else -1
        
        if idx_id_est == -1 or idx_voce == -1: return mappa_colori
        
        themes = {0: 'FFFFFF', 1: '000000', 2: 'E7E6E6', 3: '44546A', 4: '4472C4', 
                  5: 'ED7D31', 6: 'A5A5A5', 7: 'FFC000', 8: '5B9BD5', 9: '70AD47'}
                  
        for r in range(2, ws.max_row + 1):
            id_est = str(ws.cell(row=r, column=idx_id_est+1).value).strip()
            c = ws.cell(row=r, column=idx_voce+1)
            
            if id_est and c.fill and c.fill.start_color and id_est not in mappa_colori:
                if c.fill.start_color.type == 'rgb':
                    rgb = c.fill.start_color.rgb
                    if isinstance(rgb, str) and rgb != '00000000':
                        mappa_colori[id_est] = "#" + (rgb[2:] if len(rgb) == 8 else rgb)
                elif c.fill.start_color.type == 'theme':
                    t = c.fill.start_color.theme
                    if t in themes: 
                        mappa_colori[id_est] = "#" + themes[t]
    except Exception:
        pass
    return mappa_colori

# --- 4. MOTORE PRINCIPALE ---
def elabora_calcoli(df, aliquota_v_loc, codice_stato, mappa_colori, m_prefix, original_filename):
    # Setup Etichette
    label_v_loc = f"{codice_stato} ({int(aliquota_v_loc * 100)}%)"
    label_vies = "VIES (0%)"

    # Mappatura Colonne
    col_names = {col.lower().strip(): col for col in df.columns}
    col_id_estratto = col_names.get('id estratto')
    col_netto_iva = col_names.get('netto iva')
    col_valore_iva = col_names.get('valore iva')
    col_valore = col_names.get('valore estratto pdf')
    col_nome = col_names.get('nome')
    col_id = col_names.get('id') 
    
    if not all([col_id_estratto, col_netto_iva, col_valore, col_nome]):
        st.error("⚠️ Errore: Nel file mancano colonne fondamentali (es. 'ID Estratto', 'Netto Iva', 'Valore Estratto PDF', 'Nome').")
        return

    # Estrazione Periodo di Riferimento
    periodo_riferimento = "Non specificato"
    if col_id:
        row_periodo = df[df[col_id].astype(str).str.strip() == 'OrSt_O']
        if not row_periodo.empty:
            periodo_riferimento = str(row_periodo.iloc[0][col_valore]).strip()
    
    if periodo_riferimento == "Non specificato" or periodo_riferimento == 'nan':
        row_periodo = df[df[col_id_estratto].astype(str).str.strip() == 'OrSt_O']
        if not row_periodo.empty:
            periodo_riferimento = str(row_periodo.iloc[0][col_valore]).strip()

    if periodo_riferimento and periodo_riferimento != 'nan':
        st.markdown(f"<div class='period-box'>📅 <b>Periodo di Riferimento:</b> {periodo_riferimento}</div>", unsafe_allow_html=True)

    # Pulizia e Formattazione DataFrame
    df[col_id_estratto] = df[col_id_estratto].astype(str).str.strip()
    df[col_netto_iva] = df[col_netto_iva].fillna('').astype(str).str.strip().str.upper()
    df[col_valore_iva] = df[col_valore_iva].fillna('').astype(str).str.strip().str.upper() if col_valore_iva else ''
    df[col_valore] = pd.to_numeric(df[col_valore], errors='coerce').fillna(0.0)

    def col_c(id_target): return mappa_colori.get(id_target, '#1A7EBF')

    # ==========================================
    # SEZIONE: FATTURATO ORDINI (FAT_OR)
    # ==========================================
    st.markdown("### 📊 Analisi Fatturato Ordini")
    
    df_fat_or = df[df[col_id_estratto] == 'FAT_OR']
    IM_F_T = dec(df_fat_or[df_fat_or[col_netto_iva] == 'SI'][col_valore].sum())
    VAT = dec(df_fat_or[df_fat_or[col_netto_iva] == ''][col_valore].sum())
    FAT_OR = IM_F_T + VAT

    IM_F_VAT = (VAT / aliquota_v_loc).quantize(Decimal('0.01'), ROUND_HALF_UP) if aliquota_v_loc > 0 else Decimal('0.00')
    IM_F_VIES = IM_F_T - IM_F_VAT

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"<div class='metric-box' style='border-left-color: {col_c('FAT_OR')};'><div class='metric-title'>FATTURATO TOTALE (FAT_OR)</div><div class='metric-value' style='color: {col_c('FAT_OR')};'>€ {FAT_OR:,.2f}</div></div>", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div class='metric-box' style='border-left-color: {col_c('FAT_OR')};'><div class='metric-title'>IMPONIBILE TOTALE (IM_F_T)</div><div class='metric-value' style='color: {col_c('FAT_OR')};'>€ {IM_F_T:,.2f}</div><div class='metric-sub'><span style='color: #8E44AD;'>■</span> Soggetto a IVA {label_v_loc}: <b>€ {IM_F_VAT:,.2f}</b><br><span style='color: #F39C12;'>■</span> Soggetto a {label_vies}: <b>€ {IM_F_VIES:,.2f}</b></div></div>", unsafe_allow_html=True)
    with col3:
        st.markdown(f"<div class='metric-box' style='border-left-color: #E74C3C;'><div class='metric-title'>IVA {label_v_loc} (VAT)</div><div class='metric-value' style='color: #E74C3C;'>€ {VAT:,.2f}</div></div>", unsafe_allow_html=True)

    # ==========================================
    # SEZIONE: RIMBORSO INVENTARIO (RI_IN_T)
    # ==========================================
    st.markdown("---")
    st.markdown("### 📦 Rimborso Inventario")
    
    df_ri_in = df[df[col_id_estratto] == 'RI_IN_T']
    RI_IN_T = dec(df_ri_in[col_valore].sum())
    
    componenti_html = ""
    for _, row in df_ri_in.iterrows():
        info_fiscale = f"Netto IVA: <b>{str(row[col_netto_iva])}</b> | Tipo IVA: <b>{str(row.get(col_valore_iva, ''))}</b>"
        componenti_html += f"<div style='margin-bottom: 5px;'><span style='color: {col_c('RI_IN_T')}; font-size: 18px;'>■</span> {str(row[col_nome])}: <b>€ {dec(row[col_valore]):,.2f}</b><br><span style='font-size:12px; color:#888; margin-left: 18px;'>({info_fiscale})</span></div>"

    col_ri1, col_ri2 = st.columns([1, 2])
    with col_ri1:
        st.markdown(f"<div class='metric-box' style='border-left-color: {col_c('RI_IN_T')};'><div class='metric-title'>RIMBORSO INVENTARIO TOT (RI_IN_T)</div><div class='metric-value' style='color: {col_c('RI_IN_T')};'>€ {RI_IN_T:,.2f}</div><div class='metric-sub'>{componenti_html}</div></div>", unsafe_allow_html=True)

    # ==========================================
    # SEZIONE: ANALISI RIMBORSI E NOTE DI CREDITO
    # ==========================================
    st.markdown("---")
    st.markdown("### 🔄 Analisi Rimborsi")

    df_ri_c_am = df[df[col_id_estratto] == 'RI_CO_AM']
    RI_C_AM_TOT = dec(df_ri_c_am[col_valore].sum())
    RI_C_AM_IMP = (RI_C_AM_TOT / Decimal('1.22')).quantize(Decimal('0.01'), ROUND_HALF_UP)
    IVA_IT_RI_C = RI_C_AM_TOT - RI_C_AM_IMP

    df_nc = df[df[col_id_estratto] == 'NC']
    IM_R_T = dec(df_nc[df_nc[col_netto_iva] == 'SI'][col_valore].sum())
    IVA_NC = dec(df_nc[df_nc[col_netto_iva] == ''][col_valore].sum())
    NC = IM_R_T + IVA_NC
    
    IMPO_R = dec(df_nc[df_nc[col_id] == 'Impo_R'][col_valore].sum()) if col_id else IVA_NC
    IM_R_VAT = (IMPO_R / aliquota_v_loc).quantize(Decimal('0.01'), ROUND_HALF_UP) if aliquota_v_loc > 0 else Decimal('0.00')
    IM_R_VIES = IM_R_T - IM_R_VAT

    RIM_T = RI_C_AM_TOT + NC
    Rimb_R = dec(df[df[col_id] == 'Rimb_R'][col_valore].sum()) if col_id else Decimal('0.00')
    check_html = f"<span style='color:#27AE60; font-weight:bold;'>✅ CHECK OK: RIM_T ({RIM_T:,.2f}) = Rimb_R ({Rimb_R:,.2f})</span>" if RIM_T == Rimb_R else f"<span style='color:#E74C3C; font-weight:bold;'>❌ CHECK FALLITO: RIM_T ({RIM_T:,.2f}) ≠ Rimb_R ({Rimb_R:,.2f})</span><br><span style='color:#E74C3C; font-size:12px;'>Differenza Rilevata: € {RIM_T - Rimb_R:,.2f}</span>"
    
    comp_ri_c_am_html = "".join([f"<span style='color: #8E44AD;'>■</span> {str(r[col_nome])}: <b>€ {dec(r[col_valore]):,.2f}</b><br>" for _, r in df_ri_c_am.iterrows()])

    col_rim1, col_rim2, col_rim3 = st.columns(3)
    with col_rim1:
        st.markdown(f"<div class='metric-box' style='border-left-color: {col_c('RIM_T')};'><div class='metric-title'>RIMBORSI TOTALE (RIM_T)</div><div class='metric-value' style='color: {col_c('RIM_T')};'>€ {RIM_T:,.2f}</div><div class='metric-sub'>{check_html}</div></div>", unsafe_allow_html=True)
    with col_rim2:
        st.markdown(f"<div class='metric-box' style='border-left-color: {col_c('RI_CO_AM')};'><div class='metric-title'>RIMBORSO COSTI AMAZON (RI_C_AM)</div><div class='metric-value' style='color: {col_c('RI_CO_AM')};'>€ {RI_C_AM_TOT:,.2f}</div><div class='metric-sub'><span style='color: #8E44AD;'>■</span> Imponibile Scorporato: <b>€ {RI_C_AM_IMP:,.2f}</b><br><span style='color: #E74C3C;'>■</span> IVA IT (22%): <b>€ {IVA_IT_RI_C:,.2f}</b><br><br><b>Dettaglio LORDO:</b><br>{comp_ri_c_am_html}</div></div>", unsafe_allow_html=True)
    with col_rim3:
        st.markdown(f"<div class='metric-box' style='border-left-color: {col_c('NC')};'><div class='metric-title'>NOTE DI CREDITO (NC)</div><div class='metric-value' style='color: {col_c('NC')};'>€ {NC:,.2f}</div><div class='metric-sub'><b>Imponibile Totale (IM_R_T): € {IM_R_T:,.2f}</b><br><hr style='margin: 5px 0;'><span style='color: #8E44AD;'>■</span> Imp. a IVA {label_v_loc} (IM_R_VAT): <b>€ {IM_R_VAT:,.2f}</b><br><span style='color: #F39C12;'>■</span> Imp. a {label_vies} (IM_R_VIES): <b>€ {IM_R_VIES:,.2f}</b><br><span style='color: #E74C3C;'>■</span> IVA NC {label_v_loc}: <b>€ {IVA_NC:,.2f}</b></div></div>", unsafe_allow_html=True)

    # ==========================================
    # SEZIONE: COSTI AMAZON (CO_AM)
    # ==========================================
    st.markdown("---")
    st.markdown("### 📉 Costi Amazon")

    df_co_am = df[df[col_id_estratto] == 'CO_AM']
    CO_AM_TOT = dec(df_co_am[col_valore].sum())
    CO_AM_IMP = (CO_AM_TOT / Decimal('1.22')).quantize(Decimal('0.01'), ROUND_HALF_UP)
    CO_AM_IVA = CO_AM_TOT - CO_AM_IMP
    
    FAT_AM_TOT = CO_AM_TOT + RI_C_AM_TOT
    
    comp_co_am_html = ""
    for _, row in df_co_am.iterrows():
        info_fiscale = f"Netto IVA: <b>{str(row[col_netto_iva])}</b> | Tipo IVA: <b>{str(row.get(col_valore_iva, ''))}</b>"
        comp_co_am_html += f"<div style='margin-bottom: 5px;'><span style='color: {col_c('CO_AM')}; font-size: 18px;'>■</span> {str(row[col_nome])} (CO_AM): <b>€ {dec(row[col_valore]):,.2f}</b><br><span style='font-size:12px; color:#888; margin-left: 18px;'>({info_fiscale})</span></div>"

    col_fat1, col_fat2 = st.columns([1, 2])
    with col_fat1:
        st.markdown(f"<div class='metric-box' style='border-left-color: {col_c('CO_AM')};'><div class='metric-title'>COSTI AMAZON (CO_AM)</div><div class='metric-value' style='color: {col_c('CO_AM')};'>€ {CO_AM_TOT:,.2f}</div><div class='metric-sub'><span style='color: #8E44AD;'>■</span> Imponibile Scorporato: <b>€ {CO_AM_IMP:,.2f}</b><br><span style='color: #E74C3C;'>■</span> IVA IT (22%): <b>€ {CO_AM_IVA:,.2f}</b><br><br><b>Dettaglio LORDE:</b><br>{comp_co_am_html}</div></div>", unsafe_allow_html=True)

    # ==========================================
    # SEZIONE: SCHEMA FINALE E QUADRATURA
    # ==========================================
    st.markdown("---")
    st.markdown("### 🏆 Schema Finale e Quadratura")

    SALDO_INIZIALE = dec(df[df[col_id] == 'SaIn_S'][col_valore].sum()) if col_id else Decimal('0.00')
    IMPORTO_RISERVA = dec(df[df[col_id] == 'ImRi_I'][col_valore].sum()) if col_id else Decimal('0.00')
    RICAVI_NETTI_AMAZON = dec(df[df[col_id] == 'RiNe_S'][col_valore].sum()) if col_id else Decimal('0.00')

    CO_T = FAT_OR + NC + RI_IN_T
    SALDO_VENDITORE = SALDO_INIZIALE + CO_T + FAT_AM_TOT 
    RICAVI_NETTI_CALCOLATI = SALDO_VENDITORE + IMPORTO_RISERVA 

    if RICAVI_NETTI_CALCOLATI == RICAVI_NETTI_AMAZON:
        badge_quadratura = f"<div style='background-color:#D5F5E3; padding:15px; border-radius:8px; border: 2px solid #27AE60; text-align:center;'><h3 style='color:#27AE60; margin:0;'>✅ QUADRATURA PERFETTA</h3><p style='color:#1E8449; margin:0; font-size:18px;'>I Ricavi Netti calcolati corrispondono al centesimo con il report Amazon.</p></div>"
    else:
        badge_quadratura = f"<div style='background-color:#FADBD8; padding:15px; border-radius:8px; border: 2px solid #E74C3C; text-align:center;'><h3 style='color:#E74C3C; margin:0;'>❌ QUADRATURA FALLITA</h3><p style='color:#C0392B; margin:0; font-size:18px;'>Differenza rilevata: <b>€ {RICAVI_NETTI_CALCOLATI - RICAVI_NETTI_AMAZON:,.2f}</b></p></div>"

    col_schema1, col_schema2 = st.columns([1, 1])
    
    with col_schema1:
        st.markdown(f"""
        <div style="background-color: white; border: 1px solid #ddd; border-radius: 8px; padding: 20px; font-size: 16px;">
            <div style="display:flex; justify-content:space-between; margin-bottom:10px;">
                <span><b>1. SALDO INIZIALE</b></span><span>€ {SALDO_INIZIALE:,.2f}</span>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:10px;">
                <span><b>2. CORRISPETTIVO TOT (CO_T)</b> <span style="font-size:12px; color:#888;">(FAT_OR + NC + RI_IN_T)</span></span><span>€ {CO_T:,.2f}</span>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:10px;">
                <span><b>3. FATTURA AMAZON (FAT_AM)</b> <span style="font-size:12px; color:#888;">(CO_AM + RI_C_AM)</span></span><span style="color:#E74C3C;">€ {FAT_AM_TOT:,.2f}</span>
            </div>
            <hr style="margin: 10px 0;">
            <div style="display:flex; justify-content:space-between; margin-bottom:10px; font-size:18px; color:#1A7EBF;">
                <span><b>SALDO VENDITORE</b></span><span><b>€ {SALDO_VENDITORE:,.2f}</b></span>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:10px;">
                <span><b>4. IMPORTO RISERVA</b></span><span style="color:#E74C3C;">€ {IMPORTO_RISERVA:,.2f}</span>
            </div>
            <hr style="border: 2px solid #5C5D5F; margin: 15px 0;">
            <div style="display:flex; justify-content:space-between; font-size:22px;">
                <span><b>RICAVI NETTI</b></span><span><b>€ {RICAVI_NETTI_CALCOLATI:,.2f}</b></span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
    with col_schema2:
        st.markdown(badge_quadratura, unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background-color: #f8f9fa; border-radius: 8px; padding: 20px; margin-top: 20px; text-align:center;">
            <p style="color:#5C5D5F; margin-bottom:5px;">Valore Ufficiale Report Amazon (RiNe_S):</p>
            <h2 style="margin:0; color:#2A85C4;">€ {RICAVI_NETTI_AMAZON:,.2f}</h2>
        </div>
        """, unsafe_allow_html=True)

    # ==========================================
    # SEZIONE: RIEPILOGO CONTABILE E DOWNLOAD
    # ==========================================
    st.markdown("---")
    st.markdown("### 💼 Riepilogo Contabile")

    IM_CO = IM_F_T + RI_IN_T + IM_R_T
    IVA_CO = VAT + IVA_NC
    
    if aliquota_v_loc > 0:
        IM_CO_VAT = (IVA_CO / aliquota_v_loc).quantize(Decimal('0.01'), ROUND_HALF_UP)
    else:
        IM_CO_VAT = Decimal('0.00')
        
    IM_CO_VIES = IM_CO - IM_CO_VAT

    # Quadratura su Corrispettivo Totale
    CHECK_COR = IM_CO + IVA_CO
    if CO_T == CHECK_COR:
        check_cor_html = f"<div style='background-color:#D5F5E3; border: 1px solid #27AE60; padding:10px; border-radius:5px; margin-bottom:15px; color:#1E8449; font-weight:bold;'>✅ QUADRATURA CORRISPETTIVO OK: Imponibile + IVA = € {CHECK_COR:,.2f}</div>"
    else:
        diff_cor = CO_T - CHECK_COR
        check_cor_html = f"<div style='background-color:#FADBD8; border: 1px solid #E74C3C; padding:10px; border-radius:5px; margin-bottom:15px; color:#C0392B; font-weight:bold;'>❌ QUADRATURA FALLITA: Differenza rilevata € {diff_cor:,.2f}</div>"

    # --- Creazione Dinamica Nomi File e Progressivo ---
    file_num = ""
    num_match = re.search(r'AM_[A-Z]{2}\s*(\d+)', original_filename, re.IGNORECASE)
    if num_match:
        file_num = f" {num_match.group(1)}"
        
    progressivo_val = f"{m_prefix}{file_num}"
    periodo_pulito = str(periodo_riferimento).replace('/', '.').replace('\\', '.')
    nome_file_export = f"{progressivo_val} - Estratto Conto Riclassificato {periodo_pulito}.xlsx"
    nome_pdf_export = nome_file_export.replace('.xlsx', '')

    col_riepilogo1, col_riepilogo2 = st.columns([2, 1])

    with col_riepilogo1:
        st.markdown(f"""
        <div style="background-color: #F4F6F7; border-left: 5px solid #8E44AD; border-radius: 8px; padding: 20px; font-size: 16px;">
            {check_cor_html}
            <div style="display:flex; justify-content:space-between; margin-bottom:10px;">
                <span><b>CORRISPETTIVO TOT (CO_T)</b> <span style="font-size:12px; color:#888;">(FAT_OR + NC + RI_IN_T)</span></span>
                <span style="font-size:22px; font-weight:bold; color:#2C3E50;">€ {CO_T:,.2f}</span>
            </div>
            <hr style="border: 1px solid #BDC3C7; margin: 15px 0;">
            <div style="display:flex; justify-content:space-between; margin-bottom:5px;">
                <span><b>Imponibile Corrispettivo (IM_CO)</b> <span style="font-size:12px; color:#888;">(IM_F_T + RI_IN_T + IM_R_T)</span></span>
                <span><b>€ {IM_CO:,.2f}</b></span>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:5px; padding-left: 20px; font-size: 14px;">
                <span style="color: #8E44AD;">■ Soggetto a IVA {label_v_loc}</span>
                <span>€ {IM_CO_VAT:,.2f}</span>
            </div>
            <div style="display:flex; justify-content:space-between; margin-bottom:15px; padding-left: 20px; font-size: 14px;">
                <span style="color: #F39C12;">■ Soggetto a {label_vies}</span>
                <span>€ {IM_CO_VIES:,.2f}</span>
            </div>
            <div style="display:flex; justify-content:space-between;">
                <span><b style="color: #E74C3C;">IVA Corrispettivo {label_v_loc} (IVA_CO)</b> <span style="font-size:12px; color:#888;">(VAT + IVA_NC)</span></span>
                <span style="color: #E74C3C;"><b>€ {IVA_CO:,.2f}</b></span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col_riepilogo2:
        st.markdown("<br>", unsafe_allow_html=True)
        st.info("⬇️ **Esporta o Stampa Estratto Conto Riclassificato**")
        
        dati_export = {
            "Nome": [
                "STATO", "PROGRESSIVO", "PERIODO", "DATA PAGAMENTO", "RICAVI NETTI", 
                "SALDO INIZIALE", "IMPORTO DI RISERVA", 
                "CORRISPETTIVO TOT", "NETTO IVA", "VIES 0%", "IMPOSTE", "FATTURA AMAZON"
            ],
            "Valore": [
                codice_stato, progressivo_val, periodo_riferimento, "", 
                float(RICAVI_NETTI_CALCOLATI), float(SALDO_INIZIALE), float(IMPORTO_RISERVA), 
                float(CO_T), float(IM_CO_VAT), float(IM_CO_VIES), float(IVA_CO), float(FAT_AM_TOT)
            ]
        }
        df_export = pd.DataFrame(dati_export)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Estratto Conto Riclassificato')
            ws = writer.sheets['Estratto Conto Riclassificato']
            
            fmt_accounting = '_ * #,##0.00 €_ ;_ * -#,##0.00 €_ ;_ * "-"?? €_ ;_ @_ '
            font_black_bold = Font(color="000000", bold=True)
            font_black_normal = Font(color="000000")
            
            color_ricavi = "27AE60"
            color_gray = "E0E0E0"
            color_corr = col_c('FAT_OR').replace('#', '')
            color_fat_am = col_c('CO_AM').replace('#', '')
            
            fill_ricavi = PatternFill(start_color=color_ricavi, end_color=color_ricavi, fill_type='solid')
            fill_gray = PatternFill(start_color=color_gray, end_color=color_gray, fill_type='solid')
            fill_corr = PatternFill(start_color=color_corr, end_color=color_corr, fill_type='solid')
            fill_fat_am = PatternFill(start_color=color_fat_am, end_color=color_fat_am, fill_type='solid')
            
            for row in range(2, 14):
                ws[f'A{row}'].font = font_black_normal
                ws[f'B{row}'].font = font_black_normal
                if row >= 6:
                    ws[f'B{row}'].number_format = fmt_accounting
            
            ws['A6'].fill = fill_ricavi; ws['A6'].font = font_black_bold
            ws['B6'].fill = fill_ricavi; ws['B6'].font = font_black_bold
            ws['A7'].fill = fill_gray; ws['A7'].font = font_black_normal
            ws['B7'].fill = fill_gray; ws['B7'].font = font_black_normal
            ws['A8'].fill = fill_gray; ws['A8'].font = font_black_normal
            ws['B8'].fill = fill_gray; ws['B8'].font = font_black_normal
            ws['A9'].fill = fill_corr; ws['A9'].font = font_black_bold
            ws['B9'].fill = fill_corr; ws['B9'].font = font_black_bold
            ws['A13'].fill = fill_fat_am; ws['A13'].font = font_black_bold
            ws['B13'].fill = fill_fat_am; ws['B13'].font = font_black_bold

        output.seek(0)
        
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="📥 Scarica Excel",
                data=output,
                file_name=nome_file_export,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        with col_dl2:
            # Aggiriamo il blocco di sicurezza di Streamlit usando un componente HTML dedicato
            html_code = f"""
            <button onclick="
                try {{
                    let old = window.parent.document.title;
                    window.parent.document.title = '{nome_pdf_export}';
                    window.parent.print();
                    setTimeout(function() {{ window.parent.document.title = old; }}, 2000);
                }} catch (e) {{
                    alert('Il tuo browser blocca la stampa automatica. Usa la scorciatoia da tastiera Ctrl+P (o Cmd+P su Mac) per stampare o salvare in PDF.');
                }}
            " style="background-color: #E74C3C; color: white; border-radius: 8px; border: none; padding: 10px 15px; font-weight: bold; width: 100%; cursor: pointer; font-family: sans-serif; font-size: 16px; box-sizing: border-box;">
                🖨️ Stampa PDF
            </button>
            """
            st.components.v1.html(html_code, height=45)

# --- 5. SEZIONE CARICAMENTO E SCELTA MARKETPLACE ---
st.markdown("---")
st.markdown("### Seleziona il Marketplace e Carica il Report")

marketplaces = {
    "Seleziona un'opzione...": None, "🇮🇹 IT - Italia": "AM_IT", "🇫🇷 FR - Francia": "AM_FR",
    "🇩🇪 DE - Germania": "AM_DE", "🇪🇸 ES - Spagna": "AM_ES", "🇳🇱 NL - Olanda": "AM_NL",
    "🇵🇱 PL - Polonia": "AM_PL", "🇸🇪 SE - Svezia": "AM_SE", "🇮🇪 IE - Irlanda": "AM_IE", "🌍 XX - Altro": "AM_XX"
}

col_stato, col_file = st.columns([1, 2])

with col_stato:
    sel_market_full = st.selectbox("Seleziona il Paese (Serve per l'Aliquota V_LOC):", list(marketplaces.keys()))
    m_prefix = marketplaces[sel_market_full]
    codice_stato = sel_market_full.split()[1] if m_prefix else "XX"

with col_file:
    file_caricato = st.file_uploader("Carica o trascina qui il file Estratto Conto Tab", type=["xlsx"])

if file_caricato:
    if m_prefix is None:
        st.error("⚠️ ATTENZIONE: Seleziona il Paese dal menù a tendina per poter applicare l'aliquota IVA corretta!")
    else:
        with st.spinner("Lettura del file ed estrazione layout in corso..."):
            try:
                mappa_colori = estrai_colori_tema(file_caricato)
                df_importato = pd.read_excel(file_caricato)
                elabora_calcoli(df_importato, ALIQUOTE_IVA.get(m_prefix, Decimal('0.00')), codice_stato, mappa_colori, m_prefix, file_caricato.name)
            except Exception as e:
                st.error(f"Errore durante la lettura del file: {e}")
